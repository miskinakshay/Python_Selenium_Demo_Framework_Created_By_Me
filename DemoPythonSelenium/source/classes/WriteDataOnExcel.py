
from selenium import webdriver
import unittest
import HtmlTestRunner
from openpyxl import Workbook
from source.classes.ReadingDataOnExcel import *

class writeDataOnExcel:

    @classmethod
    def test_try(self):

        wb = Workbook()
        sheet = wb.active
        n=4
        # list comprehension use here (It should be divide into 4 brackets....beacuse n=4)
        final=[ExcelData[i * n:(i + 1) * n] for i in range((len(ExcelData) + n - 1) // n)]
        #print(final)
        for r in final:
            sheet.append(r)

        wb.save("C:\\Users\\Administrator\\PycharmProjects\\DemoPythonSelenium\\source\\Excel_File\\demo.xlsx")


ob = writeDataOnExcel
ob.test_try()


'''# insert single data.
    def test_WriteDataOnExcel(self):

        wb = openpyxl.Workbook()
        sheet = wb.active                     #  It will select/ get active sheet on Excel file
        c1 = sheet.cell(row=1, column=1)      # writing values to cells
        c1.value = "ABCD"

        c2 = sheet.cell(row=1, column=2)
        c2.value = "XYZ"

        c3 = sheet['A2']
        c3.value = "PQRS"

        c4 = sheet['B2']
        c4.value = "XYZXYZ "

       # wb.save("C:\\Users\\Administrator\\PycharmProjects\\DemoPythonSelenium\\source\\Excel_File\\demo.xlsx")
'''
    





