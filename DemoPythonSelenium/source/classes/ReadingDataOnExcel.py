
from selenium import webdriver
import unittest
import HtmlTestRunner
import openpyxl


class readingDataOnExcel:

    def test_ReadingDataOnExcel(self):

        wb=openpyxl.load_workbook("C:\\Users\\Administrator\\PycharmProjects\\DemoPythonSelenium\\source\\Excel_File\\demo2.xlsx")
        sh=wb['Sheet1']
        rows=sh.max_row
        columns=sh.max_column
        self.result = []

        for i in range(1,rows+1):
            for j in range(1,columns+1):
                #result = sh.cell(i, j)
                #print(result.value)

                self.result.append(sh.cell(i,j).value)    # Data Store in (result)List

        #final=result
        print(self.result)
        #print(final)
        # Here to print data on List(result)

ob=readingDataOnExcel()
ob.test_ReadingDataOnExcel()
ExcelData=ob.result
print(ExcelData)